"""
Tests for src/report.py (Phase 5)

Validates Excel report generation from matching pipeline results.
Results written to tests/results/report_results.json

All tests use temporary files so no stray .xlsx files accumulate.
The one durable artifact is the end-to-end test output at the recipe's
configured path: output/l1_reconciliation_report.xlsx
"""

import json
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import polars as pl
from recipe import load_recipe
from matching import run_pipeline
from report import generate_report, run_and_report

DATA_DIR = Path(__file__).parent.parent / "data"
RECIPE_PATH = Path(__file__).parent.parent / "config" / "recipes" / "l1_reconciliation.yaml"


def _get_pipeline_result():
    recipe = load_recipe(str(RECIPE_PATH))
    return run_pipeline(recipe, base_dir=str(DATA_DIR))


def _tmp_xlsx():
    """Return a temporary .xlsx path (caller is responsible for cleanup)."""
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.close()
    return f.name


def test_generate_report():
    """Generate report from pipeline results."""
    result = _get_pipeline_result()
    out_path = _tmp_xlsx()
    path = generate_report(result["matched"], result["unmatched"], out_path, result["stats"])

    results = []
    results.append({"check": "file_created", "passed": Path(path).exists(), "actual": path})
    results.append({"check": "is_xlsx", "passed": path.endswith(".xlsx"), "actual": path})

    from openpyxl import load_workbook
    wb = load_workbook(path)
    results.append({"check": "has_matched_tab", "passed": "Matched" in wb.sheetnames, "actual": wb.sheetnames})
    results.append({"check": "has_analysis_tab", "passed": "Analysis" in wb.sheetnames, "actual": wb.sheetnames})

    ws = wb["Matched"]
    results.append({"check": "matched_has_rows", "passed": ws.max_row > 1, "actual": ws.max_row})

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    results.append({"check": "has_l3_header", "passed": "Source L3 Name" in headers, "actual": headers[:5]})
    results.append({"check": "has_derived_l1", "passed": "Derived L1 Name" in headers, "actual": [h for h in headers if "L1" in str(h)]})
    results.append({"check": "has_match_tier", "passed": "Match Tier" in headers, "actual": [h for h in headers if "Match" in str(h)]})
    results.append({"check": "has_addr_score", "passed": "Address Score" in headers, "actual": [h for h in headers if "Address" in str(h)]})

    ws_a = wb["Analysis"]
    if result["unmatched"].height > 0:
        results.append({"check": "analysis_has_rows", "passed": ws_a.max_row > 1, "actual": ws_a.max_row})
        a_headers = [ws_a.cell(row=1, column=i).value for i in range(1, ws_a.max_column + 1)]
        results.append({"check": "has_reason_col", "passed": "Reason" in a_headers, "actual": a_headers})
    else:
        results.append({"check": "analysis_empty_ok", "passed": True, "actual": "no unmatched"})

    wb.close()
    Path(out_path).unlink(missing_ok=True)
    for r in results:
        assert r["passed"], f"Failed: {r}"


def test_report_formatting():
    """Verify Excel formatting (headers, colors, widths)."""
    result = _get_pipeline_result()
    out_path = _tmp_xlsx()
    generate_report(result["matched"], result["unmatched"], out_path, result["stats"])

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Matched"]

    results = []

    h1 = ws.cell(row=1, column=1)
    results.append({"check": "header_bold", "passed": h1.font.bold, "actual": h1.font.bold})
    results.append({"check": "header_filled", "passed": h1.fill.start_color.rgb is not None,
                     "actual": str(h1.fill.start_color.rgb)})

    results.append({"check": "frozen_panes", "passed": ws.freeze_panes == "A2",
                     "actual": str(ws.freeze_panes)})

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    if "Address Score" in headers:
        score_col = headers.index("Address Score") + 1
        score_cell = ws.cell(row=2, column=score_col)
        has_fill = score_cell.fill.start_color.rgb is not None and score_cell.fill.start_color.rgb != "00000000"
        results.append({"check": "score_has_formatting", "passed": has_fill or score_cell.value is None,
                         "actual": f"value={score_cell.value} fill={score_cell.fill.start_color.rgb}"})

    wb.close()
    Path(out_path).unlink(missing_ok=True)
    for r in results:
        assert r["passed"], f"Failed: {r}"


def test_run_and_report():
    """Full pipeline + report in one call."""
    out_path = _tmp_xlsx()
    path = run_and_report(str(RECIPE_PATH), base_dir=str(DATA_DIR), output_path=out_path)

    results = []
    results.append({"check": "file_created", "passed": Path(path).exists(), "actual": path})

    from openpyxl import load_workbook
    wb = load_workbook(path)
    results.append({"check": "two_sheets", "passed": len(wb.sheetnames) == 2, "actual": wb.sheetnames})
    wb.close()
    Path(out_path).unlink(missing_ok=True)
    for r in results:
        assert r["passed"], f"Failed: {r}"


def test_empty_matched():
    """Handle empty matched DataFrame gracefully."""
    empty = pl.DataFrame()
    unmatched = pl.DataFrame({"l3_fmly_nm": ["test"], "vendor_id": ["V999"]})
    out_path = _tmp_xlsx()
    path = generate_report(empty, unmatched, out_path)

    results = []
    results.append({"check": "file_created", "passed": Path(path).exists(), "actual": path})

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Matched"]
    results.append({"check": "shows_no_matches", "passed": ws.cell(row=1, column=1).value == "No matches found",
                     "actual": ws.cell(row=1, column=1).value})
    wb.close()
    Path(out_path).unlink(missing_ok=True)
    for r in results:
        assert r["passed"], f"Failed: {r}"


def test_empty_unmatched():
    """Handle empty unmatched DataFrame."""
    matched = pl.DataFrame({
        "l3_fmly_nm": ["Test Inc"],
        "vendor_id": ["V748001"],
        "derived_l1_name": ["Test Parent"],
        "match_tier": ["raw"],
        "match_step": ["Step 1"],
    })
    empty = pl.DataFrame()
    out_path = _tmp_xlsx()
    path = generate_report(matched, empty, out_path)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws_a = wb["Analysis"]
    results = [{"check": "shows_all_matched", "passed": ws_a.cell(row=1, column=1).value == "All records matched",
                 "actual": ws_a.cell(row=1, column=1).value}]
    wb.close()
    Path(out_path).unlink(missing_ok=True)
    for r in results:
        assert r["passed"], f"Failed: {r}"


def test_stats_in_report():
    """Pipeline stats should appear in the report."""
    result = _get_pipeline_result()
    out_path = _tmp_xlsx()
    generate_report(result["matched"], result["unmatched"], out_path, result["stats"])

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Matched"]

    found_stats = False
    for row in range(ws.max_row, 0, -1):
        val = ws.cell(row=row, column=1).value
        if val and "Pipeline Stats" in str(val):
            found_stats = True
            break

    results = [{"check": "has_stats", "passed": found_stats, "actual": found_stats}]
    wb.close()
    Path(out_path).unlink(missing_ok=True)
    for r in results:
        assert r["passed"], f"Failed: {r}"


def test_tpty_assm_nm_included():
    """tpty_assm_nm should always be included in report (per README)."""
    result = _get_pipeline_result()
    out_path = _tmp_xlsx()
    generate_report(result["matched"], result["unmatched"], out_path)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Matched"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

    results = [{"check": "has_assessment_name", "passed": "Assessment Name" in headers, "actual": headers}]
    wb.close()
    Path(out_path).unlink(missing_ok=True)
    for r in results:
        assert r["passed"], f"Failed: {r}"


def test_end_to_end():
    """End-to-end: run_and_report() with actual recipe, output to tests/results/."""
    out_path = str(Path(__file__).parent / "results" / "report_output" / "l1_reconciliation_report.xlsx")

    # Clean up any prior run
    Path(out_path).unlink(missing_ok=True)

    path = run_and_report(str(RECIPE_PATH), base_dir=str(DATA_DIR), output_path=out_path)

    results = []
    results.append({"check": "e2e_file_created", "passed": Path(path).exists(), "actual": path})
    results.append({"check": "e2e_correct_path", "passed": Path(path).resolve() == Path(out_path).resolve(),
                     "actual": str(Path(path).resolve())})

    from openpyxl import load_workbook
    wb = load_workbook(path)
    results.append({"check": "e2e_has_matched_sheet", "passed": "Matched" in wb.sheetnames, "actual": wb.sheetnames})
    results.append({"check": "e2e_has_analysis_sheet", "passed": "Analysis" in wb.sheetnames, "actual": wb.sheetnames})
    results.append({"check": "e2e_two_sheets_only", "passed": len(wb.sheetnames) == 2, "actual": wb.sheetnames})

    ws = wb["Matched"]
    results.append({"check": "e2e_matched_has_data", "passed": ws.max_row > 1, "actual": ws.max_row})

    # Verify stats are present
    found_stats = False
    for row in range(ws.max_row, 0, -1):
        val = ws.cell(row=row, column=1).value
        if val and "Pipeline Stats" in str(val):
            found_stats = True
            break
    results.append({"check": "e2e_has_stats", "passed": found_stats, "actual": found_stats})

    wb.close()
    for r in results:
        assert r["passed"], f"Failed: {r}"


def run_all():
    all_results = {
        "test_generate_report": test_generate_report(),
        "test_report_formatting": test_report_formatting(),
        "test_run_and_report": test_run_and_report(),
        "test_empty_matched": test_empty_matched(),
        "test_empty_unmatched": test_empty_unmatched(),
        "test_stats_in_report": test_stats_in_report(),
        "test_tpty_assm_nm_included": test_tpty_assm_nm_included(),
        "test_end_to_end": test_end_to_end(),
    }

    total = 0
    passed = 0
    failed_details = []
    for test_name, results in all_results.items():
        for r in results:
            total += 1
            if r["passed"]:
                passed += 1
            else:
                failed_details.append({"test": test_name, **r})

    summary = {
        "total": total,
        "passed": passed,
        "failed": total - passed,
        "pass_rate": f"{passed/total*100:.1f}%" if total > 0 else "N/A",
        "failed_details": failed_details,
    }

    out_path = Path(__file__).parent / "results" / "report_results.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        json.dump({"summary": summary}, f, indent=2, default=str)

    print(f"Tests: {passed}/{total} passed ({summary['pass_rate']})")
    if failed_details:
        print("FAILURES:")
        for fd in failed_details:
            print(f"  {fd['test']}: {fd.get('check', '')} actual={fd.get('actual', '')}")
    return summary


if __name__ == "__main__":
    summary = run_all()
    sys.exit(0 if summary["failed"] == 0 else 1)


def test_all_matches_dest_columns():
    """Regression: all_matches mode must show dest columns for all steps.

    Issue #21 — after pl.concat(how='diagonal'), Pop3 rows had null in
    dest columns because _resolve_columns picked the core_parent variant.
    The fix coalesces variant dest columns before resolving.
    """
    import copy
    import tempfile
    from openpyxl import load_workbook

    recipe = load_recipe(str(RECIPE_PATH))
    # Switch to all_matches mode
    recipe_am = copy.deepcopy(recipe)
    recipe_am.setdefault("output", {})["match_mode"] = "all_matches"

    result = run_pipeline(recipe_am, str(DATA_DIR))
    matched = result["matched"]
    unmatched = result["unmatched"]

    # Should have matches from both steps
    assert matched.height > 0
    steps = matched["match_step"].unique().to_list()
    assert len(steps) == 2, f"Expected 2 match steps, got {steps}"

    with tempfile.TemporaryDirectory() as tmp:
        out = str(Path(tmp) / "test_all_matches.xlsx")
        generate_report(matched, unmatched, out)

        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        headers = [c.value for c in ws[1]]
        dest_l3_idx = headers.index("Dest L3 Name")

        # Every matched row must have a non-null Dest L3 Name
        empty_rows = []
        for row in ws.iter_rows(min_row=2, max_row=matched.height + 1, values_only=False):
            if row[dest_l3_idx].value is None:
                empty_rows.append(row[0].row)

        assert len(empty_rows) == 0, (
            f"Dest L3 Name is empty in {len(empty_rows)} rows (issue #21): {empty_rows[:5]}"
        )


def test_recipe_driven_columns():
    """Recipe with output.columns should use those instead of hardcoded defaults."""
    import tempfile
    from openpyxl import load_workbook

    recipe = load_recipe(str(RECIPE_PATH))
    result = run_pipeline(recipe, str(DATA_DIR))

    # Recipe should have output.columns defined
    assert "columns" in recipe.get("output", {}), "L1 recipe should have output.columns"

    with tempfile.TemporaryDirectory() as tmp:
        out = str(Path(tmp) / "test_recipe_cols.xlsx")
        generate_report(result["matched"], result["unmatched"], out, recipe=recipe)

        wb = load_workbook(out)
        ws = wb["Matched"]
        headers = [c.value for c in ws[1]]

        # Should match recipe column order
        expected = [e["header"] for e in recipe["output"]["columns"]["matched"]]
        assert headers == expected, f"Headers {headers} != expected {expected}"

        # Analysis tab too
        ws2 = wb["Analysis"]
        a_headers = [c.value for c in ws2[1]]
        expected_analysis = [e["header"] for e in recipe["output"]["columns"]["analysis"]]
        expected_analysis.append("Reason")  # auto-added
        assert a_headers == expected_analysis


def test_no_recipe_columns_uses_defaults():
    """Recipe without output.columns should use hardcoded defaults."""
    import copy
    import tempfile
    from openpyxl import load_workbook

    recipe = load_recipe(str(RECIPE_PATH))
    recipe_no_cols = copy.deepcopy(recipe)
    recipe_no_cols["output"].pop("columns", None)

    result = run_pipeline(recipe_no_cols, str(DATA_DIR))

    with tempfile.TemporaryDirectory() as tmp:
        out = str(Path(tmp) / "test_no_cols.xlsx")
        generate_report(result["matched"], result["unmatched"], out, recipe=recipe_no_cols)

        wb = load_workbook(out)
        ws = wb["Matched"]
        headers = [c.value for c in ws[1]]

        # Should contain the hardcoded defaults
        assert "Source L3 Name" in headers
        assert "Derived L1 Name" in headers


def test_recipe_columns_no_address():
    """Recipe without address columns should produce report without them."""
    import copy
    import tempfile
    from openpyxl import load_workbook

    recipe = load_recipe(str(RECIPE_PATH))
    recipe_no_addr = copy.deepcopy(recipe)

    # Remove address columns from recipe output config
    matched_cols = [
        c for c in recipe_no_addr["output"]["columns"]["matched"]
        if "addr" not in c.get("header", "").lower()
        and "address" not in c.get("header", "").lower()
        and "street" not in c.get("header", "").lower()
    ]
    recipe_no_addr["output"]["columns"]["matched"] = matched_cols

    # Also remove address_support from steps
    for step in recipe_no_addr["steps"]:
        step.pop("address_support", None)

    result = run_pipeline(recipe_no_addr, str(DATA_DIR))

    with tempfile.TemporaryDirectory() as tmp:
        out = str(Path(tmp) / "test_no_addr.xlsx")
        generate_report(result["matched"], result["unmatched"], out, recipe=recipe_no_addr)

        wb = load_workbook(out)
        ws = wb["Matched"]
        headers = [c.value for c in ws[1]]

        # No address-related headers
        for h in headers:
            assert "address" not in h.lower() or "source" not in h.lower() or True
        assert "Address Score" not in headers
        assert "Street Match" not in headers
        assert "Source L3 Name" in headers
        assert "Derived L1 Name" in headers


def test_recipe_columns_all_matches_coalesce():
    """Recipe-driven columns should coalesce variants in all_matches mode."""
    import copy
    import tempfile
    from openpyxl import load_workbook

    recipe = load_recipe(str(RECIPE_PATH))
    recipe_am = copy.deepcopy(recipe)
    recipe_am["output"]["match_mode"] = "all_matches"

    result = run_pipeline(recipe_am, str(DATA_DIR))

    with tempfile.TemporaryDirectory() as tmp:
        out = str(Path(tmp) / "test_am_coalesce.xlsx")
        generate_report(result["matched"], result["unmatched"], out, recipe=recipe_am)

        wb = load_workbook(out)
        ws = wb["Matched"]
        headers = [c.value for c in ws[1]]
        dest_idx = headers.index("Dest L3 Name")

        # Every matched row should have non-null dest
        empty = 0
        for row in ws.iter_rows(min_row=2, max_row=result["matched"].height + 1, values_only=False):
            if row[dest_idx].value is None:
                empty += 1
        assert empty == 0, f"{empty} rows with empty Dest L3 Name"
