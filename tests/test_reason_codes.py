"""
Tests for Issue #32: enriched analysis tab with reason codes.

Validates that unmatched records get appropriate reason codes:
- no_name_match: no name match found in any destination
- addr_below_threshold: name matched but address score below threshold
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import copy
import polars as pl
from matching import run_pipeline
from recipe import load_recipe

RECIPE_PATH = Path(__file__).resolve().parent.parent / "config" / "recipes" / "l1_reconciliation.yaml"
DATA_DIR = Path(__file__).resolve().parent.parent / "data"


def test_unmatched_have_reason_code():
    """All unmatched records should have a reason_code column."""
    r = load_recipe(str(RECIPE_PATH))
    res = run_pipeline(r, base_dir=str(DATA_DIR))
    u = res["unmatched"]

    assert "reason_code" in u.columns
    assert "rejection_step" in u.columns
    assert "best_rejected_score" in u.columns

    # Every row should have a non-null reason
    assert u["reason_code"].null_count() == 0


def test_no_name_match_reason():
    """Records that don't match any destination name get no_name_match."""
    r = load_recipe(str(RECIPE_PATH))
    res = run_pipeline(r, base_dir=str(DATA_DIR))
    u = res["unmatched"]

    no_match = u.filter(pl.col("reason_code") == "no_name_match")
    # With default thresholds, all unmatched should be no_name_match
    assert no_match.height == u.height
    # These should have no rejection_step or score
    assert no_match["rejection_step"].null_count() == no_match.height


def test_addr_below_threshold_reason():
    """Records that name-match but fail address threshold get addr_below_threshold."""
    r = copy.deepcopy(load_recipe(str(RECIPE_PATH)))
    # Set very high threshold to force some address rejections
    for step in r["steps"]:
        if "address_support" in step:
            step["address_support"]["threshold"] = 99

    res = run_pipeline(r, base_dir=str(DATA_DIR))
    u = res["unmatched"]

    addr_rej = u.filter(pl.col("reason_code") == "addr_below_threshold")
    assert addr_rej.height >= 1, "Expected at least one addr_below_threshold rejection"

    # These should have a rejection_step and score
    assert addr_rej["rejection_step"].null_count() == 0
    assert addr_rej["best_rejected_score"].null_count() == 0

    # Score should be below the threshold
    for score in addr_rej["best_rejected_score"].to_list():
        assert score < 99, f"Rejected score {score} should be below threshold 99"


def test_reason_codes_in_report():
    """Reason codes should appear in the Analysis tab of the Excel report."""
    import tempfile
    from openpyxl import load_workbook
    from report import generate_report

    r = load_recipe(str(RECIPE_PATH))
    res = run_pipeline(r, base_dir=str(DATA_DIR))

    with tempfile.TemporaryDirectory() as tmp:
        out = str(Path(tmp) / "test_reasons.xlsx")
        generate_report(res["matched"], res["unmatched"], out, res["stats"], recipe=r)

        wb = load_workbook(out)
        ws = wb["Analysis"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

        assert "Reason Code" in headers, f"Expected 'Reason Code' in headers: {headers}"

        # Check that reason values are written
        reason_col = headers.index("Reason Code") + 1
        reasons = [ws.cell(row=i, column=reason_col).value for i in range(2, ws.max_row + 1)]
        assert all(r is not None for r in reasons), f"Some reason cells are empty: {reasons}"

        wb.close()


def test_matched_plus_unmatched_still_equals_total():
    """Reason code enrichment should not affect the matched+unmatched invariant."""
    r = copy.deepcopy(load_recipe(str(RECIPE_PATH)))
    for step in r["steps"]:
        if "address_support" in step:
            step["address_support"]["threshold"] = 99

    res = run_pipeline(r, base_dir=str(DATA_DIR))
    s = res["stats"]
    assert s["matched_count"] + s["unmatched_count"] == s["total_source"]
