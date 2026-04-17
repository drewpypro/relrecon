"""Tests for run summary generation (Issue #53)."""

import polars as pl
import pytest
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from summary import generate_summary, generate_mermaid, write_summary_tab


# --- Fixtures ---

def _recipe():
    """Minimal recipe for testing."""
    return {
        "name": "Test Recipe",
        "description": "A test reconciliation",
        "sources": {
            "ref": {"file": "ref.csv", "type": "trusted_reference"},
            "data": {"file": "data.csv", "type": "multi_population"},
        },
        "populations": {
            "target": {
                "source": "data",
                "record_key": "id",
                "filter": [
                    {"field": "id", "op": "starts_with", "value": "V"},
                ],
            },
            "junk": {
                "source": "data",
                "filter": [
                    {"field": "status", "op": "eq", "value": "deleted"},
                    {"field": "user", "op": "contains_any", "values": ["bot1", "bot2"], "join": "and"},
                ],
                "action": "exclude",
            },
        },
        "steps": [
            {
                "name": "Exact to ref",
                "source": "target",
                "destination": "ref",
                "match_fields": [{"source": "name", "destination": "name", "method": "exact", "tiers": ["raw"]}],
                "address_support": {"source": ["addr"], "destination": ["addr"], "threshold": 60},
                "date_gate": {"field": "updated", "max_age_years": 2, "applies_to": "destination"},
            },
            {
                "name": "Fuzzy to ref",
                "source": "target",
                "destination": "ref",
                "match_fields": [{"source": "name", "destination": "name", "method": "fuzzy", "threshold": 70, "tiers": ["raw"]}],
                "address_support": {"source": ["addr"], "destination": ["addr"], "threshold": 50},
                "filters": [{"field": "updated", "op": "max_age_years", "value": 2, "applies_to": "destination"}],
            },
        ],
        "output": {"format": "xlsx"},
    }


def _stats():
    return {"total_source": 100, "matched_count": 85, "unmatched_count": 15}


def _matched_df():
    return pl.DataFrame({
        "match_step": ["Exact to ref"] * 70 + ["Fuzzy to ref"] * 15,
        "name": [f"vendor_{i}" for i in range(85)],
    })


# --- Markdown ---

class TestGenerateSummary:
    def test_contains_recipe_name(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "Test Recipe" in md

    def test_contains_description(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "A test reconciliation" in md

    def test_contains_population_info(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "100 records" in md
        assert "data.csv" in md
        assert 'starts with "V"' in md
        assert "matching target" in md

    def test_contains_excluded_population(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "junk" in md
        assert "excluded" in md
        assert "deleted" in md

    def test_contains_destination_population(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        # ref is a source (not a population), so destinations that are
        # populations should be labeled
        # In this test recipe, destination "ref" is a source, not a population
        # so this just checks the population section exists
        assert "Populations" in md

    def test_contains_match_counts(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "85 of 100 (85%)" in md
        assert "15 (see Analysis tab)" in md

    def test_per_step_counts(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "| 70 |" in md  # exact
        assert "| 15 |" in md  # fuzzy

    def test_step_thresholds(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        # Exact step: name threshold is "-", fuzzy step: "70"
        assert "| - |" in md  # exact name threshold
        assert "| 70 |" in md  # fuzzy name threshold
        assert "| 60 |" in md  # addr threshold step 1
        assert "| 50 |" in md  # addr threshold step 2

    def test_date_filter_from_date_gate(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "Updated < 2yr" in md

    def test_date_filter_from_filters(self):
        """Step 2 uses filters instead of date_gate."""
        md = generate_summary(_recipe(), _stats(), _matched_df())
        lines = md.split("\n")
        step_lines = [l for l in lines if l.startswith("| ")]
        date_mentions = [l for l in step_lines if "< 2yr" in l]
        assert len(date_mentions) >= 2

    def test_cascade_explanation(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "move to the next" in md

    def test_timing_included(self):
        timing = {"load": 0.01, "setup": 0.00, "match": 0.05, "resolve": 0.01}
        md = generate_summary(_recipe(), _stats(), _matched_df(), timing=timing)
        assert "Timing" in md
        assert "Load" in md
        assert "Total" in md

    def test_timing_omitted_when_none(self):
        md = generate_summary(_recipe(), _stats(), _matched_df(), timing=None)
        assert "Timing" not in md

    def test_no_description(self):
        recipe = _recipe()
        del recipe["description"]
        md = generate_summary(recipe, _stats(), _matched_df())
        assert "Test Recipe" in md  # still works

    def test_zero_records(self):
        stats = {"total_source": 0, "matched_count": 0, "unmatched_count": 0}
        md = generate_summary(_recipe(), stats, pl.DataFrame({"match_step": []}))
        assert "0 of 0" in md


# --- Excel ---

class TestWriteSummaryTab:
    def test_creates_content(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        write_summary_tab(ws, _recipe(), _stats(), _matched_df())
        # Check title row
        assert "Test Recipe" in ws.cell(row=1, column=1).value
        # Check stats are present
        values = [ws.cell(row=r, column=2).value or "" for r in range(1, 20)]
        all_text = " ".join(str(v) for v in values)
        assert "85 of 100" in all_text
        assert "15" in all_text

    def test_step_rows_present(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        write_summary_tab(ws, _recipe(), _stats(), _matched_df())
        # Find header row
        all_values = []
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            all_values.append(row)
        header_rows = [r for r in all_values if r and r[0] == "Step"]
        assert len(header_rows) == 1


# --- Mermaid ---

class TestGenerateMermaid:
    def test_starts_with_flowchart(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        assert md.startswith("flowchart TD")

    def test_contains_source_population(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        assert "target: 100 records" in md

    def test_contains_step_nodes(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        assert "Step 1: Exact to ref" in md
        assert "Step 2: Fuzzy to ref" in md

    def test_contains_matched_count(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        assert "Matched: 85" in md

    def test_contains_unmatched_count(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        assert "Unmatched: 15" in md

    def test_per_step_counts(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        assert "70 matched" in md
        assert "15 matched" in md

    def test_cascade_remaining(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        # After step 1 (70 matched), 30 remaining
        assert "30 remaining" in md

    def test_dashed_cascade_lines(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        assert "-.->" in md

    def test_no_duplicate_connections(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df())
        lines = md.split("\n")
        # Should only have one connection from Pop
        pop_connections = [l for l in lines if l.strip().startswith("Pop -->")]
        assert len(pop_connections) == 1

    def test_detailed_mode(self):
        md = generate_mermaid(_recipe(), _stats(), _matched_df(), detailed=True)
        assert "addr >=" in md
        assert "updated" in md

    def test_empty_steps(self):
        recipe = _recipe()
        recipe["steps"] = []
        md = generate_mermaid(recipe, _stats(), _matched_df())
        assert md == ""

    def test_mermaid_in_summary_markdown(self):
        md = generate_summary(_recipe(), _stats(), _matched_df())
        assert "```mermaid" in md
        assert "flowchart TD" in md
