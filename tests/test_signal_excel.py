"""Tests for signal analysis Excel report generation and new analyses."""

import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import polars as pl
from signal_analysis import (
    analyze_dataset, top_ngrams, singleton_tokens, near_duplicate_tokens,
    token_position_frequency, token_length_distribution, numeric_token_ratio,
)
from signal_excel import generate_signal_excel


def _load_multi():
    """Load multi-pop test dataset."""
    return pl.read_csv(str(Path(__file__).parent.parent / "data" / "tp_multi_pop_dataset.csv"))


def _run_analysis(columns=None):
    """Run signal analysis and return results."""
    df = _load_multi()
    cols = columns or ["l3_fmly_nm", "hq_addr1"]
    return analyze_dataset(df, cols, unicode_mode="profile_only")


# ---------------------------------------------------------------------------
# Ngram tests
# ---------------------------------------------------------------------------

def test_top_bigrams_raw():
    """Bigrams at raw tier should return multi-word tokens."""
    df = _load_multi()
    bigrams = top_ngrams(df["l3_fmly_nm"], tier="raw", n_gram=2, n=10)
    assert len(bigrams) > 0
    for gram, count in bigrams:
        assert " " in gram, f"Expected space in bigram: {gram}"
        assert count > 0


def test_top_trigrams_raw():
    """Trigrams at raw tier should return three-word tokens."""
    df = _load_multi()
    trigrams = top_ngrams(df["l3_fmly_nm"], tier="raw", n_gram=3, n=10)
    assert len(trigrams) > 0
    for gram, count in trigrams:
        parts = gram.split(" ")
        assert len(parts) == 3, f"Expected 3 words in trigram: {gram}"


def test_top_bigrams_clean():
    """Bigrams at clean tier should be lowercase."""
    df = _load_multi()
    bigrams = top_ngrams(df["l3_fmly_nm"], tier="clean", n_gram=2, n=10)
    assert len(bigrams) > 0
    for gram, count in bigrams:
        assert gram == gram.lower(), f"Expected lowercase bigram: {gram}"


def test_top_ngrams_invalid_tier():
    """Ngrams with invalid tier should raise ValueError."""
    df = _load_multi()
    try:
        top_ngrams(df["l3_fmly_nm"], tier="normalized", n_gram=2)
        assert False, "Expected ValueError"
    except ValueError:
        pass


def test_analyze_column_has_ngrams():
    """analyze_column results should include bigram and trigram data."""
    results = _run_analysis(["l3_fmly_nm"])
    col_data = results["columns"]["l3_fmly_nm"]
    for key in ["bigrams_raw", "bigrams_clean", "trigrams_raw", "trigrams_clean"]:
        assert key in col_data, f"Missing {key} in column analysis"
        assert len(col_data[key]) > 0, f"Empty {key}"


# ---------------------------------------------------------------------------
# Singleton tests
# ---------------------------------------------------------------------------

def test_singleton_tokens_finds_unique():
    """Singletons should find tokens appearing exactly once."""
    df = _load_multi()
    singletons = singleton_tokens(df["l3_fmly_nm"], tier="clean", n=50)
    assert len(singletons) > 0
    for token, count in singletons:
        assert count == 1, f"Singleton {token} has count {count}"


def test_singleton_tokens_empty_series():
    """Singletons on empty series should return empty list."""
    empty = pl.Series("empty", [], dtype=pl.String)
    assert singleton_tokens(empty) == []


def test_analyze_column_has_singletons():
    """analyze_column should include singletons."""
    results = _run_analysis(["l3_fmly_nm"])
    assert "singletons" in results["columns"]["l3_fmly_nm"]


# ---------------------------------------------------------------------------
# Near-duplicate tests
# ---------------------------------------------------------------------------

def test_near_duplicates_structure():
    """Near-duplicates should return properly structured results."""
    df = _load_multi()
    dupes = near_duplicate_tokens(df["l3_fmly_nm"], tier="clean")
    # May or may not find any in synthetic data, but structure should be correct
    for d in dupes:
        assert "token1" in d
        assert "token2" in d
        assert "similarity" in d
        assert d["similarity"] >= 85


def test_near_duplicates_empty_series():
    """Near-duplicates on empty series should return empty list."""
    empty = pl.Series("empty", [], dtype=pl.String)
    assert near_duplicate_tokens(empty) == []


def test_analyze_column_has_near_duplicates():
    """analyze_column should include near_duplicates."""
    results = _run_analysis(["l3_fmly_nm"])
    assert "near_duplicates" in results["columns"]["l3_fmly_nm"]


# ---------------------------------------------------------------------------
# Token position tests
# ---------------------------------------------------------------------------

def test_token_position_frequency_structure():
    """Token positions should have first, last, middle keys."""
    df = _load_multi()
    pos = token_position_frequency(df["l3_fmly_nm"], tier="clean")
    assert "first" in pos
    assert "last" in pos
    assert "middle" in pos
    # Should have some data
    assert len(pos["first"]) > 0
    assert len(pos["last"]) > 0


def test_token_position_empty_series():
    """Token positions on empty series should return empty dicts."""
    empty = pl.Series("empty", [], dtype=pl.String)
    pos = token_position_frequency(empty)
    assert pos == {"first": [], "last": [], "middle": []}


def test_analyze_column_has_positions():
    """analyze_column should include token_positions."""
    results = _run_analysis(["l3_fmly_nm"])
    assert "token_positions" in results["columns"]["l3_fmly_nm"]


# ---------------------------------------------------------------------------
# Token length tests
# ---------------------------------------------------------------------------

def test_token_length_distribution_structure():
    """Token length distribution should have stats and histogram."""
    df = _load_multi()
    tl = token_length_distribution(df["l3_fmly_nm"])
    assert "min" in tl
    assert "max" in tl
    assert "mean" in tl
    assert "median" in tl
    assert "histogram" in tl
    assert tl["min"] > 0
    assert tl["max"] >= tl["min"]
    assert len(tl["histogram"]) > 0


def test_token_length_empty_series():
    """Token length on empty series should return zeros."""
    empty = pl.Series("empty", [], dtype=pl.String)
    tl = token_length_distribution(empty)
    assert tl["min"] == 0
    assert tl["histogram"] == []


def test_analyze_column_has_lengths():
    """analyze_column should include token_lengths."""
    results = _run_analysis(["l3_fmly_nm"])
    assert "token_lengths" in results["columns"]["l3_fmly_nm"]


# ---------------------------------------------------------------------------
# Numeric ratio tests
# ---------------------------------------------------------------------------

def test_numeric_ratio_name_column():
    """Name column should have low numeric ratio."""
    df = _load_multi()
    nr = numeric_token_ratio(df["l3_fmly_nm"])
    assert nr["total_tokens"] > 0
    assert nr["alpha"] > nr["numeric"]  # Names should be mostly alpha


def test_numeric_ratio_id_column():
    """ID column should have high numeric ratio."""
    df = _load_multi()
    nr = numeric_token_ratio(df["vendor_id"])
    # vendor_id is alphanumeric IDs like V7001 -- mostly mixed
    assert nr["total_tokens"] > 0


def test_numeric_ratio_empty_series():
    """Numeric ratio on empty series should return zeros."""
    empty = pl.Series("empty", [], dtype=pl.String)
    nr = numeric_token_ratio(empty)
    assert nr["total_tokens"] == 0
    assert nr["numeric_pct"] == 0.0


def test_analyze_column_has_numeric_ratio():
    """analyze_column should include numeric_ratio."""
    results = _run_analysis(["l3_fmly_nm"])
    assert "numeric_ratio" in results["columns"]["l3_fmly_nm"]


# ---------------------------------------------------------------------------
# Data quality includes numeric_token_pct
# ---------------------------------------------------------------------------

def test_data_quality_has_numeric_pct():
    """Data quality should include numeric_token_pct."""
    results = _run_analysis(["l3_fmly_nm"])
    q = results["data_quality"]["l3_fmly_nm"]
    assert "numeric_token_pct" in q


# ---------------------------------------------------------------------------
# Excel generation tests
# ---------------------------------------------------------------------------

EXPECTED_SHEETS = ["Summary", "TopTokens", "Alias", "NearDuplicates",
                   "TokenProfile", "Unicode"]


def test_generate_excel_creates_file(tmp_path):
    """Excel file should be created with correct sheets."""
    results = _run_analysis()
    out = str(tmp_path / "test_output.xlsx")
    path = generate_signal_excel(results, out)
    assert Path(path).exists()

    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_summary_sheet_structure(tmp_path):
    """Summary sheet should have headers and data in expected positions."""
    results = _run_analysis()
    out = str(tmp_path / "test_summary.xlsx")
    generate_signal_excel(results, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]

    # Section headers
    assert ws["A1"].value == "Top Tokens"
    assert ws["G1"].value == "Data Quality"

    # Column headers at row 4
    assert ws["A4"].value == "colName"
    assert ws["B4"].value == "token"
    assert ws["G4"].value == "colName"
    assert ws["H4"].value == "detectedType"

    # Should have data
    assert ws["A5"].value is not None
    # Numeric % column should exist
    assert ws["M4"].value == "numeric%"


def test_top_tokens_sheet_has_all_signal_types(tmp_path):
    """TopTokens sheet should have topk, bigram, trigram and singleton rows."""
    results = _run_analysis(["l3_fmly_nm"])
    out = str(tmp_path / "test_tokens.xlsx")
    generate_signal_excel(results, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["TopTokens"]

    signal_types = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1]:
            signal_types.add(row[1])

    assert "topk" in signal_types
    assert "bigram" in signal_types
    assert "trigram" in signal_types
    assert "singleton" in signal_types


def test_alias_sheet_has_data(tmp_path):
    """Alias sheet should have variant groups from synthetic data."""
    results = _run_analysis(["l3_fmly_nm"])
    out = str(tmp_path / "test_alias.xlsx")
    generate_signal_excel(results, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["Alias"]

    assert ws["A1"].value == "canonical"
    assert ws.max_row > 1


def test_near_duplicates_sheet_exists(tmp_path):
    """NearDuplicates sheet should exist with proper headers."""
    results = _run_analysis(["l3_fmly_nm"])
    out = str(tmp_path / "test_dupes.xlsx")
    generate_signal_excel(results, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["NearDuplicates"]
    assert ws["A1"].value == "columnName"
    assert ws["B1"].value == "token1"
    assert ws["D1"].value == "similarity%"


def test_token_profile_sheet_has_sections(tmp_path):
    """TokenProfile sheet should have position, length and numeric sections."""
    results = _run_analysis(["l3_fmly_nm"])
    out = str(tmp_path / "test_profile.xlsx")
    generate_signal_excel(results, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["TokenProfile"]

    # Check section headers exist somewhere
    values = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
        if row[0]:
            values.append(row[0])

    assert any("Position" in str(v) for v in values)
    assert any("Length" in str(v) for v in values)
    assert any("Numeric" in str(v) for v in values)


def test_unicode_sheet_has_data(tmp_path):
    """Unicode sheet should have character range profiles."""
    results = _run_analysis(["l3_fmly_nm"])
    out = str(tmp_path / "test_unicode.xlsx")
    generate_signal_excel(results, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["Unicode"]

    assert ws["A1"].value == "columnName"
    assert ws.max_row > 1


def test_top_n_limits_output(tmp_path):
    """top_n parameter should limit rows in detail sheets."""
    results = _run_analysis(["l3_fmly_nm"])
    out_full = str(tmp_path / "full.xlsx")
    out_limited = str(tmp_path / "limited.xlsx")

    generate_signal_excel(results, out_full, top_n=None)
    generate_signal_excel(results, out_limited, top_n=5)

    import openpyxl
    wb_full = openpyxl.load_workbook(out_full)
    wb_limited = openpyxl.load_workbook(out_limited)

    full_rows = wb_full["TopTokens"].max_row
    limited_rows = wb_limited["TopTokens"].max_row

    assert limited_rows < full_rows


def test_empty_results(tmp_path):
    """Should handle empty analysis results gracefully."""
    results = {
        "data_quality": {},
        "columns": {},
        "_raw_series": {},
        "aggregated_stopwords": {},
        "aggregated_aliases": {},
    }
    out = str(tmp_path / "empty.xlsx")
    path = generate_signal_excel(results, out)
    assert Path(path).exists()

    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == EXPECTED_SHEETS


# ---------------------------------------------------------------------------
# CLI integration tests
# ---------------------------------------------------------------------------

def test_cli_xlsx_format():
    """CLI --signal-format xlsx should produce Excel output."""
    result = subprocess.run(
        [sys.executable, "-m", "src", "--analyze", "data/tp_multi_pop_dataset.csv",
         "--columns", "l3_fmly_nm", "--signal-format", "xlsx",
         "--signal-output", "output/cli_test.xlsx", "--top", "10"],
        capture_output=True, text=True,
        cwd=str(Path(__file__).parent.parent),
    )
    assert result.returncode == 0, f"stderr: {result.stderr}"
    assert "Excel report saved" in result.stdout
    assert Path("output/cli_test.xlsx").exists() or \
           Path(Path(__file__).parent.parent / "output" / "cli_test.xlsx").exists()


def test_cli_both_format():
    """CLI --signal-format both should produce Excel and markdown."""
    result = subprocess.run(
        [sys.executable, "-m", "src", "--analyze", "data/tp_multi_pop_dataset.csv",
         "--columns", "l3_fmly_nm", "--signal-format", "both",
         "--signal-output", "output/cli_both_test", "--top", "5"],
        capture_output=True, text=True,
        cwd=str(Path(__file__).parent.parent),
    )
    assert result.returncode == 0, f"stderr: {result.stderr}"
    assert "Excel report saved" in result.stdout
    assert "Markdown report saved" in result.stdout


def test_cli_default_format_is_md():
    """CLI without --signal-format should default to markdown (stdout)."""
    result = subprocess.run(
        [sys.executable, "-m", "src", "--analyze", "data/tp_multi_pop_dataset.csv",
         "--columns", "l3_fmly_nm", "--top", "3"],
        capture_output=True, text=True,
        cwd=str(Path(__file__).parent.parent),
    )
    assert result.returncode == 0
    assert "Signal Analysis Report" in result.stdout


def test_cli_md_includes_new_sections():
    """CLI markdown output should include new analysis sections."""
    result = subprocess.run(
        [sys.executable, "-m", "src", "--analyze", "data/tp_multi_pop_dataset.csv",
         "--columns", "l3_fmly_nm", "--top", "5"],
        capture_output=True, text=True,
        cwd=str(Path(__file__).parent.parent),
    )
    assert result.returncode == 0
    out = result.stdout
    assert "Singleton tokens" in out
    assert "Token position frequency" in out or "position" in out.lower()
    assert "Token length stats" in out
    assert "Token types" in out  # numeric ratio
