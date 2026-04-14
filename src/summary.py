"""
Run summary generator.

Produces a plain-English summary of a pipeline run by combining the
recipe config (what was configured) with the pipeline stats (what
actually happened). Output as markdown string or Excel "Summary" tab.
"""

import polars as pl


def _describe_filters(pop_cfg: dict) -> str:
    """Describe population filters in plain English."""
    filters = pop_cfg.get("filter", [])
    if not filters:
        return "everything remaining"

    parts = []
    for f in filters:
        field = f.get("field", "?")
        op = f.get("op", "?")
        val = f.get("value", f.get("values", "?"))

        if op == "starts_with":
            parts.append(f'{field} starts with "{val}"')
        elif op == "not_starts_with":
            parts.append(f'{field} does not start with "{val}"')
        elif op == "eq":
            parts.append(f'{field} is "{val}"')
        elif op == "neq":
            parts.append(f'{field} is not "{val}"')
        elif op == "contains":
            parts.append(f'{field} contains "{val}"')
        elif op == "contains_any":
            quoted = [f'"{v}"' for v in val] if isinstance(val, list) else [f'"{val}"']
            parts.append(f'{field} contains any of {", ".join(quoted)}')
        else:
            parts.append(f"{field} {op} {val}")

    join_mode = "and"
    for f in filters:
        if "join" in f:
            join_mode = f["join"]
            break

    return f" {join_mode} ".join(parts)


def _describe_step(step: dict, matched_count: int) -> dict:
    """Extract step info into a flat dict for rendering."""
    mf = step.get("match_fields", [{}])[0]
    method = mf.get("method", "?")
    threshold = mf.get("threshold", 100 if method == "exact" else "?")

    addr = step.get("address_support", {})
    addr_threshold = addr.get("threshold", "none")

    # Date filter from date_gate or filters
    date_desc = "none"
    dg = step.get("date_gate")
    if dg:
        date_desc = f'{dg["field"]} within {dg["max_age_years"]} years'
    else:
        for f in step.get("filters", []):
            if f.get("op") == "max_age_years":
                date_desc = f'{f["field"]} within {f["value"]} years'
                break

    return {
        "name": step.get("name", "?"),
        "destination": step.get("destination", "?"),
        "method": method.capitalize(),
        "name_threshold": f"{threshold}%" if isinstance(threshold, int) else threshold,
        "addr_threshold": f"\u2265{addr_threshold}%" if isinstance(addr_threshold, (int, float)) else addr_threshold,
        "date_filter": date_desc,
        "matched": matched_count,
    }


def _format_timing(timing: dict) -> str:
    """Format timing dict into a readable string."""
    phases = [("load", "Load"), ("setup", "Setup"), ("match", "Match"), ("resolve", "Resolve")]
    parts = [f"{label} {timing[k]:.2f}s" for k, label in phases if k in timing]
    total = sum(timing.get(k, 0) for k, _ in phases)
    parts.append(f"Total {total:.2f}s")
    return " | ".join(parts)


def generate_summary(recipe: dict, stats: dict, matched_df: pl.DataFrame,
                     timing: dict | None = None) -> str:
    """Generate a markdown run summary from recipe config + pipeline stats.

    Args:
        recipe: The parsed recipe dict
        stats: Pipeline stats dict (total_source, matched_count, unmatched_count)
        matched_df: The matched DataFrame (for per-step counts)
        timing: Optional pipeline timing dict (load, setup, match, resolve)

    Returns:
        Markdown string
    """
    name = recipe.get("name", "Unnamed Recipe")
    desc = recipe.get("description", "")
    total = stats.get("total_source", 0)
    matched = stats.get("matched_count", 0)
    unmatched = stats.get("unmatched_count", 0)
    pct = round(matched / total * 100) if total > 0 else 0

    # Per-step counts from match_step column
    step_counts = {}
    if matched_df is not None and "match_step" in matched_df.columns:
        for row in matched_df.group_by("match_step").len().iter_rows():
            step_counts[row[0]] = row[1]

    lines = []
    lines.append(f"# {name} -- Run Summary")
    if desc:
        lines.append(f"\n*{desc}*")
    lines.append("")

    # --- Populations ---
    step_sources = {step["source"] for step in recipe.get("steps", [])}
    step_dests = {step["destination"] for step in recipe.get("steps", [])}

    lines.append("**Populations:**")
    for pop_name, pop_cfg in recipe.get("populations", {}).items():
        src_name = pop_cfg.get("source", "")
        source_file = ""
        if src_name and src_name in recipe.get("sources", {}):
            source_file = recipe["sources"][src_name].get("file", "")

        filter_desc = _describe_filters(pop_cfg)
        file_part = f" from {source_file}" if source_file else ""

        if pop_cfg.get("action") == "exclude":
            lines.append(f"- **{pop_name}:** excluded ({filter_desc})")
        elif pop_name in step_sources:
            lines.append(f"- **{pop_name}:** {total} records{file_part} ({filter_desc}) -- *matching target*")
        elif pop_name in step_dests:
            lines.append(f"- **{pop_name}:**{file_part} ({filter_desc}) -- *destination*")
        else:
            lines.append(f"- **{pop_name}:**{file_part} ({filter_desc})")
    lines.append("")

    # --- Results (trailing two spaces for markdown line breaks) ---
    lines.append(f"**Matched:** {matched} of {total} ({pct}%)  ")
    lines.append(f"**Unmatched:** {unmatched} (see Analysis tab)  ")
    if timing:
        lines.append(f"**Timing:** {_format_timing(timing)}  ")
    lines.append("")

    # --- Step table ---
    lines.append("**Matching steps (in priority order):**")
    lines.append("")
    lines.append("| Step | Against | Method | Name threshold | Address threshold | Date filter | Matched |")
    lines.append("|---|---|---|---|---|---|---|")

    for i, step in enumerate(recipe.get("steps", []), 1):
        count = step_counts.get(step.get("name", ""), 0)
        info = _describe_step(step, count)
        lines.append(
            f"| {i} | {info['destination']} | {info['method']} "
            f"| {info['name_threshold']} | {info['addr_threshold']} "
            f"| {info['date_filter']} | {info['matched']} |"
        )

    lines.append("")
    lines.append(
        "Records that don't match or fail a threshold in one step "
        "move to the next. A record is only unmatched if it fails all steps."
    )

    return "\n".join(lines)


def write_summary_tab(ws, recipe: dict, stats: dict, matched_df: pl.DataFrame,
                      timing: dict | None = None) -> None:
    """Write a Summary tab to an openpyxl worksheet.

    Args:
        ws: openpyxl Worksheet (already created)
        recipe: The parsed recipe dict
        stats: Pipeline stats dict
        matched_df: The matched DataFrame (for per-step counts)
        timing: Optional pipeline timing dict
    """
    from openpyxl.styles import Font, Alignment, PatternFill

    bold = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    name = recipe.get("name", "Unnamed Recipe")
    desc = recipe.get("description", "")
    total = stats.get("total_source", 0)
    matched = stats.get("matched_count", 0)
    unmatched = stats.get("unmatched_count", 0)
    pct = round(matched / total * 100) if total > 0 else 0

    # Per-step counts
    step_counts = {}
    if matched_df is not None and "match_step" in matched_df.columns:
        for row in matched_df.group_by("match_step").len().iter_rows():
            step_counts[row[0]] = row[1]

    # Determine population roles
    step_sources = {step["source"] for step in recipe.get("steps", [])}
    step_dests = {step["destination"] for step in recipe.get("steps", [])}

    row = 1

    # Title
    ws.cell(row=row, column=1, value=f"{name} -- Run Summary").font = bold
    row += 1
    if desc:
        ws.cell(row=row, column=1, value=desc)
        row += 1
    row += 1

    # Populations
    ws.cell(row=row, column=1, value="Populations:").font = Font(bold=True)
    row += 1
    for pop_name, pop_cfg in recipe.get("populations", {}).items():
        src_name = pop_cfg.get("source", "")
        source_file = ""
        if src_name and src_name in recipe.get("sources", {}):
            source_file = recipe["sources"][src_name].get("file", "")

        filter_desc = _describe_filters(pop_cfg)
        file_part = f" from {source_file}" if source_file else ""

        if pop_cfg.get("action") == "exclude":
            label = f"{pop_name}: excluded ({filter_desc})"
        elif pop_name in step_sources:
            label = f"{pop_name}: {total} records{file_part} ({filter_desc}) -- matching target"
        elif pop_name in step_dests:
            label = f"{pop_name}:{file_part} ({filter_desc}) -- destination"
        else:
            label = f"{pop_name}:{file_part} ({filter_desc})"
        ws.cell(row=row, column=1, value=label)
        row += 1
    row += 1

    # Results
    result_rows = [
        ("Matched", f"{matched} of {total} ({pct}%)"),
        ("Unmatched", f"{unmatched} (see Analysis tab)"),
    ]
    if timing:
        result_rows.append(("Timing", _format_timing(timing)))

    for label, value in result_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Step table
    ws.cell(row=row, column=1, value="Matching steps (in priority order):").font = Font(bold=True)
    row += 1

    headers = ["Step", "Against", "Method", "Name threshold", "Address threshold", "Date filter", "Matched"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
    row += 1

    for i, step in enumerate(recipe.get("steps", []), 1):
        count = step_counts.get(step.get("name", ""), 0)
        info = _describe_step(step, count)
        values = [
            i, info["destination"], info["method"],
            info["name_threshold"], info["addr_threshold"],
            info["date_filter"], info["matched"],
        ]
        for ci, v in enumerate(values, 1):
            ws.cell(row=row, column=ci, value=v)
        row += 1

    row += 1
    note = ws.cell(
        row=row, column=1,
        value="Records that don't match or fail a threshold in one step "
              "move to the next. A record is only unmatched if it fails all steps.",
    )
    note.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)

    # Column widths
    widths = [16, 20, 12, 18, 18, 28, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w
